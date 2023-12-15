import openai
import os
import time
import pandas as pd
import tempfile
from openpyxl import load_workbook

#以下代码以课程《中外名曲赏析》为例
workbook = load_workbook(r"raw_data\中外名曲赏析.xlsx")
sheet = workbook['Sheet1']

# 循环遍历每一行
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    value_ = row[7].value
    k = 9
    #with open(r'prompt_temp.txt', 'a', encoding='utf-8') as file:
    #    file.write('\n\n')
    #    file.write(value)
    #    file.write('\n\n')
    #    file.write('***')
    with open(r'prompt_temp.txt', 'r', encoding='utf-8') as file:
        template = file.read()
    with tempfile.NamedTemporaryFile(mode='w+',encoding='utf-8') as temp_file:
        temp_file.write(template)
        temp_file.write('\n\n')
        temp_file.write(value_)
        temp_file.write('\n\n')
        temp_file.write('***')
        temp_file.seek(0)
        content = temp_file.read()
    print(content)
    message = [
        {
            "role": "user",
            "content": content,
        },
    ]
    key_value = os.getenv('KEY_NAME') #使用自己的api_key
    result = openai.ChatCompletion.create(
        model = 'gpt-3.5-turbo-0301',
        messages = message,
        api_key = key_value,
        max_tokens = 100,
    )
    generated_text = result.choices[0].message['content']
    print(f"{generated_text}")
    split_strings = generated_text.split('\n')
    last_chars = [substring[-1] for substring in split_strings if substring]
    print(f"{last_chars}")
    for tuple in last_chars:
        if(tuple.isdigit()):
            print(f"{tuple}")
            sheet.cell(row[0].row, k,tuple)
            k += 1
    time.sleep(20)
# 保存工作簿
workbook.save(r"raw_data\中外名曲赏析.xlsx")